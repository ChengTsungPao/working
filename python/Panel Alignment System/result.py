import cv2
import numpy as np
import json
from glob import glob

from openpyxl import load_workbook

def compare_image():

    index = 0

    path1 = ".\Test Image_20210913\M3mm_Deg1_Bri255\cal_{}_L.png".format(index)
    path2 = ".\Test Image_20210913\M3mm_Deg2.5_Bri255\cal_{}_L.png".format(index)

    image1 = cv2.imread(path1, 0)
    image2 = cv2.imread(path2, 0)

    shape = np.shape(image1)
    diff = np.sum(image1 - image2) / (shape[0] * shape[1])

    print(diff)
    print(image1 - image2)


def cal_difference_distance():

    def cal_difference(path, index):
        start, end = index

        f = open(path + "result.txt", "r")
        content = f.readlines()
        content = "".join(content)

        L_table, R_table = [], []
        for index in range(start, end + 1):
            str_ = content.split("cal_{}_{}: ".format(index, "L"))[-1].split("\n")[0]
            x = int(str_.split(",")[0].split("x = ")[-1])
            y = int(str_.split(", y = ")[-1])
            L_table += [(x, y)]

        for index in range(start, end + 1):
            str_ = content.split("cal_{}_{}: ".format(index, "R"))[-1].split("\n")[0]
            x = int(str_.split(",")[0].split("x = ")[-1])
            y = int(str_.split(", y = ")[-1])
            R_table += [(x, y)]

        # answer
        L_answer_table, R_answer_table = [], []
        for index in range(start, end + 1):
            f = open(path + "cal_{}_L.json".format(index), "r")
            data = json.load(f)

            x = int(data["shapes"][3]["points"][0][0])
            y = int(data["shapes"][3]["points"][0][1])
            L_answer_table += [(x, y)]

        for index in range(start, end + 1):
            f = open(path + "cal_{}_R.json".format(index), "r")
            data = json.load(f)

            x = int(data["shapes"][3]["points"][0][0]) - 200
            y = int(data["shapes"][3]["points"][0][1])
            R_answer_table += [(x, y)]

        # diff
        diff_L, diff_R = [], []
        # print("left")
        for index in range(len(L_answer_table)):
            x, y = L_table[index]
            _x, _y = L_answer_table[index]
            # print((x, y), (_x, _y))
            # diff_L.append(((x - _x) ** 2 + (y - _y) ** 2) ** 0.5)
            diff_L.append([abs(x - _x), abs(y - _y)])

        # print("right")
        for index in range(len(R_answer_table)):
            x, y = R_table[index]
            _x, _y = R_answer_table[index]
            # print((x, y), (_x, _y))
            # diff_R.append(((x - _x) ** 2 + (y - _y) ** 2) ** 0.5)
            diff_R.append([abs(x - _x), abs(y - _y)])

        
        # print(diff_L)
        # print(diff_R)
        return diff_L, diff_R


    paths = glob(".\Test Image_20210913\M3mm_Deg*")
    # paths = [".\Test Image_20210913\M3mm_Deg2.5_Bri100\\"]

    for path in paths:
        path += "\\"
        # diff_L_translation, diff_R_translation = cal_difference(path, [0, 8])
        # diff_translation = np.mean([diff_L_translation, diff_R_translation])
        # diff_L_rotation, diff_R_rotation = cal_difference(path, [9, 10])
        # diff_rotation = np.mean([diff_L_rotation, diff_R_rotation])
        # print(path)
        # print("diff_translation", round(diff_L_translation, 2), round(diff_R_translation, 2))
        # print("diff_translation", round(diff_translation, 2))
        # print("diff_rotation", round(diff_L_rotation, 2), round(diff_R_rotation, 2))
        # print("diff_rotation", round(diff_rotation, 2))

        print(path)
        diff_L_translation, diff_R_translation = cal_difference(path, [0, 8])
        diff_translation = np.mean(diff_L_translation + diff_R_translation, axis = 0)
        diff_L_rotation, diff_R_rotation = cal_difference(path, [9, 10])
        diff_rotation = np.mean(diff_L_rotation + diff_R_rotation, axis = 0)
        # print("diff_translation", np.round(diff_L_translation, 2), np.round(diff_R_translation, 2))
        print("diff_translation", np.round(diff_translation, 2))
        # print("diff_rotation", np.round(diff_L_rotation, 2), np.round(diff_R_rotation, 2))
        print("diff_rotation", np.round(diff_rotation, 2))
        print("=================================")


def create_sheet_result():
    path = "./Test Image_20210913/"
    workbook = load_workbook(path + "result.xlsx")

    detection = {
        "L": {
            "x": ["B{}".format(index) for index in range(4, 14 + 1)],
            "y": ["C{}".format(index) for index in range(4, 14 + 1)]
        },
        "R": {
            "x": ["D{}".format(index) for index in range(4, 14 + 1)],
            "y": ["E{}".format(index) for index in range(4, 14 + 1)]
        }
    }

    groundTruth = {
        "L": {
            "x": ["F{}".format(index) for index in range(4, 14 + 1)],
            "y": ["G{}".format(index) for index in range(4, 14 + 1)]
        },
        "R": {
            "x": ["H{}".format(index) for index in range(4, 14 + 1)],
            "y": ["I{}".format(index) for index in range(4, 14 + 1)]
        }
    }


    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        degree = sheetname.split("deg")[0]
        light = sheetname.split("b")[-1]

        folder = "M3mm_Deg{}_Bri{}/".format(degree, light)
        print(folder)

        # Detection
        f = open(path + folder + "result.txt", "r")
        content = f.readlines()
        content = "".join(content)

        for index in range(10 + 1):
            str_ = content.split("cal_{}_{}: ".format(index, "L"))[-1].split("\n")[0]

            x = int(str_.split(",")[0].split("x = ")[-1])
            y = int(str_.split(", y = ")[-1])

            sheet[detection["L"]["x"][index]] = x
            sheet[detection["L"]["y"][index]] = y

        for index in range(10 + 1):
            str_ = content.split("cal_{}_{}: ".format(index, "R"))[-1].split("\n")[0]

            x = int(str_.split(",")[0].split("x = ")[-1])
            y = int(str_.split(", y = ")[-1])

            sheet[detection["R"]["x"][index]] = x
            sheet[detection["R"]["y"][index]] = y

        # Ground Truth
        for index in range(10 + 1):
            f = open(path + folder + "cal_{}_L.json".format(index), "r")
            data = json.load(f)

            x = int(data["shapes"][3]["points"][0][0])
            y = int(data["shapes"][3]["points"][0][1])

            sheet[groundTruth["L"]["x"][index]] = x
            sheet[groundTruth["L"]["y"][index]] = y

        for index in range(10 + 1):
            f = open(path + folder + "cal_{}_R.json".format(index), "r")
            data = json.load(f)

            x = int(data["shapes"][3]["points"][0][0]) - 200
            y = int(data["shapes"][3]["points"][0][1])

            sheet[groundTruth["R"]["x"][index]] = x
            sheet[groundTruth["R"]["y"][index]] = y

        workbook.save(path + "result.xlsx")


if __name__ == "__main__":
    cal_difference_distance()
    # create_sheet_result()


    